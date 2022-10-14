from importlib.resources import path
from flask import Flask, render_template, request, send_file
from werkzeug.utils import secure_filename
from werkzeug.datastructures import FileStorage
from io import BytesIO
import re
from openpyxl import load_workbook
import fitz
from form_types import *
import json

app = Flask(__name__)
app.debug = True

route = ".\\ROW_QC_Parser\\"

matches = ""
results_found = []
excel_path = ""
parsed = {}
parcels_list = {}


def search_for_text(lines, search_args):
    """
    Search for the search string within the document lines
    """
    i = 0
    for search_str in search_args:
        search_str = search_str.replace("\n", " ")
        search_str = re.escape(search_str)
        search_str = search_str.replace("\ ", "[\s\\n]+")
        print(' ')
        print(search_str) 
        results = re.findall(search_str, lines)
        results_found[i] += len(results)
        # In case multiple matches within one line
        for result in results:
            yield result
        
        i += 1


def highlight_matching_data(page, matched_values):
    """
    Highlight matching values
    """
    matches_found = 0
    for val in matched_values:
        matches_found += 1
        matching_val_area = page.search_for(val)
        highlight = page.add_highlight_annot(matching_val_area)
        highlight.update()
    return matches_found


def process_data(input_file, output_file, search_terms):
    """
    Process the pages of the PDF File
    """
    pdfDoc = fitz.open(stream=input_file.read())
    # parse_summary_page(pdfDoc)
    parse_RW53(pdfDoc)
    output_buffer = BytesIO()
    total_matches = 0
    search_args = search_terms
    for pg in range(pdfDoc.page_count):
        page = pdfDoc[pg]
        page_lines = page.get_text("text")
        matched_values = search_for_text(page_lines, search_args)
        matches_found = highlight_matching_data(page, matched_values)
        total_matches += matches_found
    global matches
    matches = f"{total_matches} Match(es) Found"
    add_results_to_worksheet(results_found)
    pdfDoc.save(output_buffer)
    pdfDoc.close()
    filename = secure_filename(input_file.filename)
    output_file.save(route + filename)
    with open(route + filename, mode='wb') as f:
        file = FileStorage(f)
        file.write(output_buffer.getbuffer())


def remove_highlight(input_file, output_file):
    pdfDoc = fitz.open(stream=input_file.read(), filetype="pdf")
    output_buffer = BytesIO()
    annot_found = 0
    for pg in range(pdfDoc.page_count):
        page = pdfDoc[pg]
        annot = page.first_annot
        while annot:
            annot_found += 1
            page.delete_annot(annot)
            annot = annot.next

    if annot_found >= 0:
        print(f"{annot_found} Annotation(s) Found In The Input File: {input_file}")
    pdfDoc.save(output_buffer)
    pdfDoc.close()
    with open(output_file, mode='wb') as f:
        f.write(output_buffer.getbuffer())


def extract_search_strs_from_excel(path):
    global excel_path
    excel_path = path
    wb = load_workbook(filename = path)
    ws = wb['Sheet1']
    values = []
    i = 2
    while ws[f"B{i}"].value is not None:
        values.append(str(ws[f"B{i}"].value))
        results_found.append(0)
        i += 1

    return values

def add_results_to_worksheet(results):
    wb = load_workbook(filename = excel_path)
    ws = wb['Sheet1']
    i = 2
    for result in results:
        ws[f"C{i}"] = result
        i += 1
    
    wb.save(excel_path)


def process_file(**kwargs):
    """
    Add/remove highlights from a single PDF File
    """
    input_file = kwargs.get('input_file')
    output_file = kwargs.get('output_file')
    if output_file is None:
        output_file = input_file
    excel_file = kwargs.get('excel')

    search_terms = extract_search_strs_from_excel(excel_file)
    action = kwargs.get('action')
    if action == "Remove":
        remove_highlight(input_file=input_file,
                        output_file=output_file)
    else:
        process_data(input_file=input_file, output_file=output_file,
                     search_terms=search_terms)


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "GET":
        return render_template("main_page.html", matches=matches, summary=parsed)

    EXCEL = request.files['excel']
    PDF = request.files['pdf']
    process_file(
        input_file=PDF, 
        output_file=PDF,
        excel=EXCEL
    )

    # return render_template("main_page.html", matches=matches, summary=parsed, parcels=parcels_list)
    return send_file(PDF.filename, as_attachment=True)

def parse_RW53(pdfDoc):
    print("parsing RW53")
    global parsed
    toc = pdfDoc.get_toc()
    for pg in range(pdfDoc.page_count):
        get_form(pdfDoc[pg])
    summary = RW53(toc[0][1], toc[0][2])
    page = pdfDoc[summary.page_numbers - 1].get_text("text")
    summary.pin = re.search('(?<=PIN: )(\d*)', page).group(0)
    summary.project_number = re.search('(?<=Project No. )(\S*)', page).group(0)
    summary.project_name = re.search('(?<=Project Name: )(.*)', page).group(0)
    summary.region = re.search('(?<=Region: )(\d*)', page).group(0)
    summary.county = re.search('(?<=County: )(.*)', page).group(0)
    summary.routes = re.search('(?<=Route\(s\): )(.*)', page).group(0)
    prepared_by = re.search('(?<=Prepared by: \()(.*)\d', page).group(0)
    prepared_by = prepared_by.replace(" ", "").split(',')
    summary.prepared_by = prepared_by[0]
    # summary.prepared_by_firm = prepared_by[1]
    summary.prepared_by_date = prepared_by[2]
    items = page.split('\n')
    parcels = []
    i = 23
    while i < len(items) - 1:
        parcel = Parcel()
        parcel.parcel_number = items[i]
        parcel.grantor = items[i+ 1]
        parcel.square_feet = items[i + 2]
        parcel.acres = items[i + 3]
        parcel.deed_type = items[i + 4]
        while (items[i + 5].startswith('RW')):
            if parcel.map_sheets == None:
                parcel.map_sheets = items[i + 5]
            else:
                parcel.map_sheets += items[i + 5]
            i += 1
        i += 5    
        parcels.append(parcel)
    # summary.parcels = parcels
    global parcels_list
    parcels_list = parcels
    parsed = summary

def parse_summary_page(pdfDoc):
    global parsed
    toc = pdfDoc.get_toc()
    summary = SummaryPage(toc[0][1], toc[0][2])
    page = pdfDoc[summary.page_numbers - 1].get_text("text")
    summary.pin = re.search('(?<=PIN: )(\d*)', page).group(0)
    summary.project_number = re.search('(?<=Project No. )(\S*)', page).group(0)
    summary.project_name= re.search('(?<=Project Name: )(.*)', page).group(0)
    summary.region = re.search('(?<=Region: )(\d*)', page).group(0)
    summary.county = re.search('(?<=County: )(.*)', page).group(0)
    summary.routes= re.search('(?<=Route\(s\): )(.*)', page).group(0)
    items = page.split('\n')
    parcels = []
    i = 23
    while i < len(items) - 1:
        parcel = Parcel()
        parcel.parcel_number = items[i]
        parcel.grantor = items[i+ 1]
        parcel.square_feet = items[i + 2]
        parcel.acres = items[i + 3]
        parcel.deed_type = items[i + 4]
        while (items[i + 5].startswith('RW')):
            if parcel.map_sheets == None:
                parcel.map_sheets = items[i + 5]
            else:
                parcel.map_sheets += items[i + 5]
            i += 1
        i += 5    
        parcels.append(parcel)
    # summary.parcels = parcels
        parsed = json.dumps(summary.__dict__)

WARRANTY_DEED = {"RW-01", "RW-03", "RW-04", "RW-22", "RW24"}
SPECIAL_WARRANTY_DEED = {"RW-02"}
QUIT_CLAIM_DEED = {"RW-05", "RW-07", "RW-08", "RW-023", "RW-28"}
PERPETUAL_EASEMENT = {"RW-09"}
TEMPORARY_CONSTRUCTION_EASEMENT = {"RW-09"}
PUBLIC_UTILITY_EASEMENT = {"RW-09"}
RIGHT_OF_WAY_DEED = {"RW-17"}
RELINQUISHMENT_OF_ACCESS_RIGHTS = {"RW-11", "RW-13"}
DEED_PLOT = {"DeedPlot"}
MAP_CHECK = {"Segment"}

def parse_rw(page):
    print("RW53")

def parse_ownership(page):
    print("ownership")

def parse_warranty_deed(page):
    print("warranty deed")

def get_form(obj):
    page = obj.get_text("text")
    page2 = obj.get_text("xml")
    if "RW53" in page:
        return parse_rw(page)
    if "RW-51" in page:
        return parse_ownership(page)
    for i in WARRANTY_DEED:
        if i in page:
            return print("Warranty Deed")
    for i in SPECIAL_WARRANTY_DEED:
        if i in page:
            return print("Special Warranty Deed")
    for i in QUIT_CLAIM_DEED:
        if i in page:
            return print("Quit Claim Deed")
    for i in PERPETUAL_EASEMENT:
        if i in page:
            return print("Perpetual Easement")
    for i in TEMPORARY_CONSTRUCTION_EASEMENT:
        if i in page:
            return print("Temporary Construction Easement")
    for i in PUBLIC_UTILITY_EASEMENT:
        if i in page:
            return print("Public Utility Easement")
    for i in RIGHT_OF_WAY_DEED:
        if i in page:
            return print("Right of Way Deed")
    for i in RELINQUISHMENT_OF_ACCESS_RIGHTS:
        if i in page:
            return print("Relinquishment of Access Rights")
    for i in DEED_PLOT:
        if i in page:
            return print("Deed Plot")
    for i in MAP_CHECK:
        if i in page:
            return print("Map Check")
    else:
        print(obj.number,  " image")    

    